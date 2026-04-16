try:
    import pandas as pd
except ImportError:
    pd = None  # pandas indisponível no servidor compartilhado
import os
from typing import Dict, List, Any, Optional
from app.core.config import settings

class FileService:
    """
    Serviço central de integração com Planilhas (Leitura e Auditoria/Append).
    """
    
    def __init__(self):
        self.input_path = settings.DATA_INPUT_PATH
        self.output_path = settings.DATA_OUTPUT_PATH
        
        # Criar diretórios se não existirem
        os.makedirs(self.input_path, exist_ok=True)
        os.makedirs(self.output_path, exist_ok=True)

    def get_categories(self) -> List[str]:
        """
        Lista categorias baseadas nos arquivos presentes na pasta de entrada.
        Ex: 'CADASTRO_CORTINAS.xlsx' -> Categoria 'CORTINAS'
        """
        files = os.listdir(self.input_path)
        categories = []
        for f in files:
            if f.endswith(('.xlsx', '.csv')) and 'CADASTRO_' in f.upper():
                name = f.upper().replace('CADASTRO_', '').split('.')[0]
                categories.append(name)
        return categories

    def read_category_data(self, category: str) -> Dict[str, Any]:
        """
        Lê um arquivo de categoria (ex: CORTINAS) e retorna todas as abas/dados.
        """
        category = category.upper()
        # Tentar encontrar o arquivo (pode ser .xlsx ou .csv)
        base_name = f"CADASTRO_{category}"
        file_path = None
        
        for ext in ['.xlsx', '.csv']:
            path = os.path.join(self.input_path, f"{base_name}{ext}")
            if os.path.exists(path):
                file_path = path
                break
        
        if not file_path:
            return {"error": f"Tabela para {category} não encontrada em {self.input_path}"}

        try:
            if file_path.endswith('.xlsx'):
                # Lê todas as abas do Excel
                excel_file = pd.ExcelFile(file_path)
                data = {}
                for sheet_name in excel_file.sheet_names:
                    df = excel_file.parse(sheet_name)
                    data[sheet_name] = df.to_dict(orient='records')
                return {"type": "excel", "sheets": data}
            else:
                # Lê CSV único
                df = pd.read_csv(file_path)
                return {"type": "csv", "data": df.to_dict(orient='records')}
        except Exception as e:
            return {"error": f"Erro ao ler arquivo: {str(e)}"}

    def append_to_audit_csv(self, filename: str, data: List[Dict[str, Any]]):
        """
        Faz o Append-only (Adiciona no final) nos CSVs de auditoria (VENDA, PRODUTOS, etc).
        Implementa a regra de negócio Industrial: Nunca deletar, sempre adicionar.
        """
        path = os.path.join(self.output_path, f"{filename}.csv")
        df_new = pd.DataFrame(data)
        
        # Se arquivo não existe, cria com header. Se existe, adiciona sem header.
        header = not os.path.exists(path)
        
        # Usar modo 'a' (append)
        df_new.to_csv(path, mode='a', index=False, header=header, encoding='utf-8-sig')

    def find_client_by_cpf(self, cpf: str) -> Optional[Dict[str, Any]]:
        """
        Busca o último registro de um cliente pelo CPF no CSV de auditoria.
        """
        path = os.path.join(self.output_path, "VENDAS_CLIENTES.csv")
        if not os.path.exists(path):
            return None
        
        try:
            df = pd.read_csv(path)
            # Pegar o último registro (mais recente) do CPF
            client_data = df[df['cpf'].astype(str) == str(cpf)].tail(1)
            if not client_data.empty:
                return client_data.iloc[0].to_dict()
        except:
            pass
        return None

    def append_order_to_excel(self, order_data: dict, output_filename: str = "Dados PED.xlsx"):
        """
        Faz o Append de um pedido industrial completo no arquivo Excel de saída.
        Escreve nas 4 abas: VENDA, VENDEDORES, VENDAS_CLIENTES, PRODUTOS.
        """
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        import os
        
        path = os.path.join(self.output_path, output_filename)
        
        if not os.path.exists(path):
            raise FileNotFoundError(f"Arquivo de saída não encontrado: {path}")
        
        try:
            wb = load_workbook(path)
            nr_ped = order_data['nr_ped']
            
            # 1. Escrever na aba VENDA
            ws_venda = wb['VENDA']
            venda_row = [
                nr_ped,
                order_data['data'],
                order_data.get('arquiteto'),
                order_data['cliente']['cliente'],
                order_data['perc_venda_comiss'],
                order_data['cliente'].get('bairro'),
                order_data['cliente'].get('cidade_uf'),
                order_data.get('vlr_av', 0),  # Calculado
                order_data.get('vlr_liquido', 0),  # Calculado
                order_data['f_pagto'],
                order_data.get('recebim'),
                order_data['qt_parc'],
                order_data.get('marketing_perc'),
                order_data.get('tx_desloc'),
                order_data.get('vlr_real'),  # Calculado
                order_data.get('comissao'),  # Calculado
                order_data.get('marketing_rs'),  # Calculado
                order_data.get('desc_acresc_perc'),
                order_data.get('desc_acresc_rs'),  # Calculado
            ]
            ws_venda.append(venda_row)
            
            # 2. Escrever na aba VENDEDORES
            ws_vendedores = wb['VENDEDORES']
            for vendedor in order_data['vendedores']:
                vendedor_row = [
                    nr_ped,
                    vendedor['vendedor'],
                    vendedor['percentual']
                ]
                ws_vendedores.append(vendedor_row)
            
            # 3. Escrever na aba VENDAS_CLIENTES
            ws_clientes = wb['VENDAS_CLIENTES']
            cliente = order_data['cliente']
            cliente_row = [
                nr_ped,
                cliente['cliente'],
                cliente['end_entrega'],
                cliente.get('bairro'),
                cliente.get('cidade_uf'),
                cliente.get('cpf'),
                cliente.get('cep'),
                cliente.get('fones'),
                cliente.get('aniversario'),
                cliente.get('obs')
            ]
            ws_clientes.append(cliente_row)
            
            # 4. Escrever na aba PRODUTOS
            ws_produtos = wb['PRODUTOS']
            for idx, produto in enumerate(order_data['produtos'], 1):
                produto_row = [
                    nr_ped,
                    produto['qt'],
                    produto['produto'],
                    produto['vlr_unit_av'],
                    produto.get('vlr_unit_real'),
                    None,  # Dt. Inst.
                    produto.get('local_instalacao'),
                    produto.get('especificacoes'),
                    produto.get('obs'),
                    produto['categoria'],
                    produto['qt'] * produto['vlr_unit_av'],  # Total à vista
                    produto.get('garantia'),
                    idx  # ITEM
                ]
                ws_produtos.append(produto_row)
            
            # Salvar arquivo
            wb.save(path)
            return {"success": True, "nr_ped": nr_ped}
            
        except Exception as e:
            raise Exception(f"Erro ao escrever no Excel: {str(e)}")
